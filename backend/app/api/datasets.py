# backend/app/api/datasets.py

from fastapi import APIRouter, UploadFile, Form, HTTPException, File, Depends
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import Optional, List, Any
import pandas as pd
from io import BytesIO
from pydantic import BaseModel
import re
import datetime
from openpyxl import load_workbook
from app.core.database import get_db, engine
from app.models.dataset import Dataset
from app.models.dataset_column import DatasetColumn
from app.models.dataset_row import DatasetRow
from app.utils.type_inference import infer_column_type
from fastapi.responses import StreamingResponse
from fastapi import Query
import json

router = APIRouter(prefix="/datasets", tags=["Datasets"])

def get_dataset_df(dataset: Dataset, db: Session) -> pd.DataFrame:
    """Helper to get DataFrame from either dynamic table or legacy DatasetRow"""
    if dataset.table_name:
        try:
            with engine.begin() as conn:
                return pd.read_sql_table(dataset.table_name, conn)
        except Exception as e:
            print(f"Error reading table {dataset.table_name}: {e}")
            return pd.DataFrame()
    else:
        rows = db.query(DatasetRow).filter_by(dataset_id=dataset.id).all()
        return pd.DataFrame([r.row_data for r in rows])

@router.get("/")
def list_datasets(db: Session = Depends(get_db)):
    datasets = db.query(Dataset).order_by(Dataset.id.desc()).all()
    return [
        {
            "id": d.id,
            "project": d.project,
            "department": d.department,
            "employeeName": d.uploaded_by,
            "fileName": d.name,
            "uploadedBy": d.uploaded_by,
            "uploadDate": d.created_at.strftime("%Y-%m-%d") if d.created_at else None,
            "fileType": d.file_type,
            "records": d.row_count,
            "status": "Completed",
            "uploadDateISO": d.created_at.isoformat() if d.created_at else None
        }
        for d in datasets
    ]

# 🔹 PREVIEW DATA (EYE BUTTON)
@router.get("/{dataset_id}/excel-view")
def get_excel_view(dataset_id: int, db: Session = Depends(get_db)):
    dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if not dataset:
        raise HTTPException(status_code=404, detail="Dataset not found")

    columns = (
        db.query(DatasetColumn)
        .filter(DatasetColumn.dataset_id == dataset_id)
        .order_by(DatasetColumn.id)
        .all()
    )

    headers = [c.column_name for c in columns]
    if dataset.table_name:
        try:
            with engine.begin() as conn:
                df = pd.read_sql_query(text(f'SELECT * FROM "{dataset.table_name}" LIMIT 1000'), conn)
        except Exception as e:
            print(f"Error reading table {dataset.table_name}: {e}")
            df = pd.DataFrame()
    else:
        rows = db.query(DatasetRow).filter_by(dataset_id=dataset_id).limit(1000).all()
        if rows:
            df = pd.DataFrame([r.row_data for r in rows])
        else:
            df = pd.DataFrame()
    
    if not df.empty:
        # If headers logic is out of sync, trust the DF
        if not headers:
            headers = df.columns.tolist()
        
        # Ensure we only try to get columns that exist in DF
        valid_headers = [h for h in headers if h in df.columns]
        # Add any new columns in DF not in headers
        extra_cols = [c for c in df.columns if c not in valid_headers]
        final_headers = valid_headers + extra_cols
        
        df = df.fillna("")
        data = df[final_headers].values.tolist()
        headers = final_headers
    else:
        data = []

    return {
        "id": dataset.id,
        "name": dataset.name,
        "type": dataset.file_type,
        "size": dataset.row_count,
        "date": dataset.created_at.strftime("%Y-%m-%d"),
        "uploadedBy": dataset.uploaded_by or "System",
        "fileData": {
            "sheets": [
                {
                    "name": "Sheet1",
                    "headers": headers,
                    "data": data
                }
            ]
        }
    }


# 🔹 DOWNLOAD DATA AGAIN
@router.get("/{dataset_id}/download")
def download_dataset(
    dataset_id: int, 
    format: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    dataset = db.query(Dataset).filter_by(id=dataset_id).first()
    if not dataset:
        return {"error": "Not found"}

    df = get_dataset_df(dataset, db)

    # Determine export format: use query param if provided, else use original file type
    export_format = format.lower() if format else dataset.file_type.lower()
    
    stream = BytesIO()
    
    if export_format == "csv":
        df.to_csv(stream, index=False)
        media_type = "text/csv"
        filename = f"{dataset.name.rsplit('.', 1)[0]}.csv"
    else:
        # Default to Excel for xlsx, xls, or any other format
        df.to_excel(stream, index=False)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"{dataset.name.rsplit('.', 1)[0]}.xlsx"

    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# 🔹 CHART DATA API
@router.get("/{dataset_id}/chart")
def get_chart_data(
    dataset_id: int,
    x: str = Query(...),
    y: str = Query(...),
    db: Session = Depends(get_db),
):
    dataset = db.query(Dataset).filter_by(id=dataset_id).first()
    if not dataset:
        return {"x": [], "y": [], "count": 0}

    x_vals = []
    y_vals = []
    
    if dataset.table_name:
        try:
            with engine.begin() as conn:
                # Need to quote the column names in case they have spaces
                query = text(f'SELECT "{x}", "{y}" FROM "{dataset.table_name}"')
                df = pd.read_sql_query(query, conn)
        except Exception as e:
            print(f"Error fetching chart data: {e}")
            df = pd.DataFrame()
    else:
        df = get_dataset_df(dataset, db)

    if not df.empty and x in df.columns and y in df.columns:
        # Filter for valid numeric Y values
        df_valid = df[pd.to_numeric(df[y], errors='coerce').notna()]
        x_vals = df_valid[x].tolist()
        y_vals = df_valid[y].astype(float).tolist()

    return {
        "x": x_vals,
        "y": y_vals,
        "count": len(x_vals)
    }

class UpdateDatasetRequest(BaseModel):
    headers: List[str]
    data: List[Any]

@router.put("/{dataset_id}/data")
def update_dataset_data(
    dataset_id: int,
    payload: UpdateDatasetRequest,
    db: Session = Depends(get_db)
):
    dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if not dataset:
        raise HTTPException(status_code=404, detail="Dataset not found")

    # Update columns
    # First, delete existing columns metadata
    db.query(DatasetColumn).filter(DatasetColumn.dataset_id == dataset_id).delete()
    
    # Add new columns metadata
    for col_name in payload.headers:
        db.add(DatasetColumn(
            dataset_id=dataset_id,
            column_name=col_name,
            data_type="string" 
        ))

    # Update rows
    if dataset.table_name:
        # Update dynamic table
        try:
            # Check if data is already list of dicts or list of lists
            if payload.data and isinstance(payload.data[0], dict):
                new_df = pd.DataFrame(payload.data)
                # Reorder to match headers and handle missing columns in some rows
                cols_to_use = [h for h in payload.headers if h in new_df.columns]
                new_df = new_df[cols_to_use]
            else:
                new_df = pd.DataFrame(payload.data, columns=payload.headers)
            
            new_df.to_sql(dataset.table_name, engine, if_exists='replace', index=False)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to update table: {e}")
    else:
        # Legacy update
        db.query(DatasetRow).filter(DatasetRow.dataset_id == dataset_id).delete()
        new_rows = []
        for row_entry in payload.data:
            if isinstance(row_entry, dict):
                row_dict = row_entry
            else:
                row_dict = {}
                for i, val in enumerate(row_entry):
                    if i < len(payload.headers):
                        row_dict[payload.headers[i]] = val
            
            new_rows.append(DatasetRow(
                dataset_id=dataset_id,
                row_data=row_dict
            ))
        db.bulk_save_objects(new_rows)
    
    # Update row count
    dataset.row_count = len(payload.data)
    
    db.commit()
    
    return {"message": "Dataset updated successfully"}

def process_excel(file_bytes: bytes):
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet = wb.active

    # 1. Handle merged cells
    merged_data = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_val = sheet.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_data[(row, col)] = top_left_val

    # 2. Extract all non-empty cells to find the bounding box
    cells_dict = {}
    min_r, max_r, min_c, max_c = float('inf'), -1, float('inf'), -1
    
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            val = merged_data.get((row, col), cell.value)
            
            if val is not None and str(val).strip() != "":
                cells_dict[(row, col)] = val
                min_r = min(min_r, row)
                max_r = max(max_r, row)
                min_c = min(min_c, col)
                max_c = max(max_c, col)

    if not cells_dict:
        return []

    # 3. Robust Header Row Detection
    header_start_row = None
    header_depth_detected = 1
    final_headers = []
    
    for r in range(min_r, min(min_r + 20, max_r + 1)):
        non_null_cells = 0
        text_cells = 0
        numeric_cells = 0
        
        for c in range(min_c, max_c + 1):
            val = cells_dict.get((r, c))
            if val is not None and str(val).strip() != "":
                non_null_cells += 1
                if isinstance(val, (int, float, complex)) and not isinstance(val, bool):
                    numeric_cells += 1
                else:
                    text_cells += 1
                    
        min_required = min(5, max_c - min_c + 1)
        if non_null_cells >= min_required:
            text_ratio = text_cells / non_null_cells
            numeric_ratio = numeric_cells / non_null_cells
            
            if text_ratio >= 0.70 and numeric_ratio <= 0.30:
                best_d = 1
                max_score = -1
                best_headers = []
                
                for d in range(1, min(5, max_r - r + 2)):
                    headers_at_d = []
                    header_block = {}
                    
                    for hr in range(r, r + d):
                        last_val = ""
                        for c in range(min_c, max_c + 1):
                            val = cells_dict.get((hr, c))
                            if val is not None and str(val).strip() != "":
                                last_val = str(val).strip()
                            
                            if hr < r + d - 1:
                                header_block[(hr, c)] = last_val
                            else:
                                header_block[(hr, c)] = str(val).strip() if val is not None else ""
                                
                    for c in range(min_c, max_c + 1):
                        col_parts = []
                        for hr in range(r, r + d):
                            part = header_block.get((hr, c), "")
                            if part and (not col_parts or col_parts[-1] != part):
                                col_parts.append(part)
                                
                        header_name = " ".join(col_parts).lower().strip()
                        header_name = re.sub(r'[.\-]', ' ', header_name)
                        header_name = re.sub(r'[^a-z0-9_\s]', '', header_name)
                        header_name = re.sub(r'\s+', '_', header_name).strip('_')
                        
                        if not header_name:
                            header_name = f"column_{c}"
                            
                        headers_at_d.append(header_name)
                        
                    unique_count = len(set(headers_at_d))
                    generic_count = sum(1 for h in headers_at_d if h.startswith("column_"))
                    
                    score = unique_count - (generic_count * 0.5)
                    
                    if score > max_score:
                        max_score = score
                        best_d = d
                        best_headers = headers_at_d
                
                has_generic = any(h.startswith("column_") for h in best_headers)
                
                if not has_generic:
                    header_start_row = r
                    header_depth_detected = best_d
                    final_headers = best_headers
                    break

    if header_start_row is None:
        raise ValueError("Could not detect a valid header row. Ensure your sheet has a header with at least 5 text columns and no generic placeholders.")
        
    header_rows_end = header_start_row + header_depth_detected - 1

    unique_headers = []
    header_counts = {}
    for h in final_headers:
        if h in header_counts:
            header_counts[h] += 1
            unique_headers.append(f"{h}_{header_counts[h]}")
        else:
            header_counts[h] = 0
            unique_headers.append(h)

    # 5. Extract rows
    data = []
    for r in range(header_rows_end + 1, max_r + 1):
        row_data = {}
        is_empty = True
        for idx, c in enumerate(range(min_c, max_c + 1)):
            val = cells_dict.get((r, c))
            
            if isinstance(val, (datetime.date, datetime.datetime)):
                val = val.isoformat()
                
            if val is not None:
                is_empty = False
                
            row_data[unique_headers[idx]] = val
            
        if not is_empty:
            data.append(row_data)

    return data

@router.post("/upload")
async def upload_dataset(
    file: UploadFile = File(...),
    industry: Optional[str] = Form(None),
    project: Optional[str] = Form(None),
    department: Optional[str] = Form(None),
    employeeName: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """
    Accepts CSV or Excel file and stores dataset, columns, and rows in DB.
    Returns dataset metadata compatible with frontend.
    """

    # 0️⃣ Check for duplicates
    # "one department cannot upload duplicate tracker(excel)"
    if department:
        existing = db.query(Dataset).filter(
            Dataset.department == department, 
            Dataset.name == file.filename
        ).first()
        if existing:
            raise HTTPException(
                status_code=400, 
                detail=f"Dataset '{file.filename}' already uploaded for department '{department}'."
            )

    # 1️⃣ Read file into DataFrame
    try:
        contents = await file.read()
        if file.filename.endswith(".csv"):
            df = pd.read_csv(BytesIO(contents))
            df = df.fillna("")
        else:
            processed_data = process_excel(contents)
            if not processed_data:
                raise HTTPException(status_code=400, detail="No readable data found.")
            df = pd.DataFrame(processed_data)
            df = df.fillna("")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read file: {e}")

    # 2️⃣ Store dataset metadata
    dataset = Dataset(
        name=file.filename,
        industry=industry,
        project=project,
        department=department,
        uploaded_by=employeeName,
        file_type=file.filename.split(".")[-1].upper(),
        row_count=len(df)
    )
    db.add(dataset)
    db.commit()
    db.refresh(dataset)

    # 3️⃣ Create Dynamic Table and Insert Data
    sanitized_name = re.sub(r'[^a-zA-Z0-9_]', '_', file.filename.split('.')[0]).lower()
    table_name = f"tracker_{dataset.id}_{sanitized_name}"[:63] # Postgres limit 63 chars

    try:
        # Create table and insert data
        df.to_sql(table_name, engine, if_exists='fail', index=False)
        
        # Update dataset with table_name
        dataset.table_name = table_name
        db.commit()
    except Exception as e:
        # Rollback metadata if table creation fails
        db.delete(dataset)
        db.commit()
        raise HTTPException(status_code=500, detail=f"Failed to create table: {e}")

    # 4️⃣ Store column metadata (Keeping for consistency/schema endpoint)
    for col in df.columns:
        db.add(DatasetColumn(
            dataset_id=dataset.id,
            column_name=col,
            data_type=infer_column_type(df[col])
        ))
    db.commit()

    # 5️⃣ Return metadata for frontend tracker
    return {
        "id": dataset.id,
        "project": dataset.project,
        "department": dataset.department,
        "employeeName": dataset.uploaded_by,
        "fileName": dataset.name,
        "uploadedBy": dataset.uploaded_by,
        "uploadDate": dataset.created_at.strftime("%Y-%m-%d"),
        "fileType": dataset.file_type,
        "records": dataset.row_count,
        "status": "Completed",
        "uploadDateISO": dataset.created_at.isoformat()
    }

# ✅ THIS IS WHERE YOUR QUESTIONED CODE GOES
@router.get("/{dataset_id}/schema")
def get_schema(dataset_id: int, db: Session = Depends(get_db)):
    return db.query(DatasetColumn).filter_by(dataset_id=dataset_id).all()


@router.get("/{dataset_id}/data")
def get_data(dataset_id: int, db: Session = Depends(get_db)):
    dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if not dataset:
        return []
    
    # For dynamic tables, query only the first 1000 rows from the database directly
    if dataset.table_name:
        try:
            with engine.begin() as conn:
                query = text(f'SELECT * FROM "{dataset.table_name}" LIMIT 1000')
                df = pd.read_sql_query(query, conn)
                return df.fillna("").to_dict(orient='records')
        except Exception as e:
            print(f"Error fetching data: {e}")
            return []

    # Fallback for legacy datasets
    rows = db.query(DatasetRow).filter_by(dataset_id=dataset_id).limit(1000).all()
    if rows:
        df = pd.DataFrame([r.row_data for r in rows])
        return df.fillna("").to_dict(orient='records')
    return []

class UpdateDatasetMetadataRequest(BaseModel):
    project: Optional[str] = None
    department: Optional[str] = None
    employeeName: Optional[str] = None

@router.put("/{dataset_id}")
def update_dataset_metadata(
    dataset_id: int,
    payload: UpdateDatasetMetadataRequest,
    db: Session = Depends(get_db)
):
    dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if not dataset:
        raise HTTPException(status_code=404, detail="Dataset not found")

    if payload.project is not None:
        dataset.project = payload.project
    if payload.department is not None:
        dataset.department = payload.department
    if payload.employeeName is not None:
        dataset.uploaded_by = payload.employeeName
    
    db.commit()
    db.refresh(dataset)
    
    return {"message": "Dataset metadata updated successfully"}

@router.delete("/{dataset_id}")
def delete_dataset(dataset_id: int, db: Session = Depends(get_db)):
    dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()

    if not dataset:
        return {"error": "Dataset not found"}

    # Drop dynamic table if exists
    if dataset.table_name:
        try:
            # Use text() for raw SQL to drop table
            db.execute(text(f'DROP TABLE IF EXISTS "{dataset.table_name}"'))
        except Exception as e:
            print(f"Error dropping table {dataset.table_name}: {e}")

    # Delete child rows first (FK safety) - for legacy data
    db.query(DatasetRow).filter(DatasetRow.dataset_id == dataset_id).delete()
    db.query(DatasetColumn).filter(DatasetColumn.dataset_id == dataset_id).delete()

    db.delete(dataset)
    db.commit()

    return {"message": "Dataset deleted successfully"}

@router.post("/{dataset_id}/process")
def process_dataset_data(
    dataset_id: int,
    db: Session = Depends(get_db)
):
    """
    Triggers re-processing of dataset data:
    1. Re-infers column types
    2. Updates column metadata
    """
    dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if not dataset:
        raise HTTPException(status_code=404, detail="Dataset not found")

    df = get_dataset_df(dataset, db)
    if df.empty:
        raise HTTPException(status_code=400, detail="No data available to process")

    # 1. Update Column Metadata (Type Inference)
    # First, delete existing columns metadata to refresh them
    db.query(DatasetColumn).filter(DatasetColumn.dataset_id == dataset_id).delete()
    
    new_columns = []
    for col in df.columns:
        inferred_type = infer_column_type(df[col])
        db.add(DatasetColumn(
            dataset_id=dataset_id,
            column_name=col,
            data_type=inferred_type
        ))
        new_columns.append({
            "column_name": col,
            "data_type": inferred_type
        })
    
    db.commit()
    
    return {
        "message": "Dataset processed and optimized successfully",
        "columns": new_columns,
        "rowCount": len(df)
    }

