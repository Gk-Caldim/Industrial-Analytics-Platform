import os
import io
import json
import datetime
import re
import time
from dotenv import load_dotenv

load_dotenv()
from contextlib import asynccontextmanager
from fastapi import FastAPI, UploadFile, File, BackgroundTasks, Depends, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import asyncpg
from openpyxl import load_workbook
import uvicorn

DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://postgres:postgres@localhost:5432/postgres")

@asynccontextmanager
async def lifespan(app: FastAPI):
    try:
        # statement_cache_size=0 is required when using transaction poolers like Supabase pgbouncer
        app.state.pool = await asyncpg.create_pool(DATABASE_URL, statement_cache_size=0)
        # Ensure registry table exists
        async with app.state.pool.acquire() as conn:
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS file_registry (
                    id SERIAL PRIMARY KEY,
                    filename TEXT NOT NULL,
                    table_name TEXT NOT NULL,
                    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
                )
            """)
        print("Database connected successfully.")
    except Exception as e:
        print(f"Failed to connect to the database: {e}")
        app.state.pool = None
    yield
    if app.state.pool:
        await app.state.pool.close()

app = FastAPI(lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def process_excel(file_bytes: bytes):
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active

    # 1. Handle merged cells
    # We create a dictionary of merged cells to easily access the top-left value
    merged_data = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_val = sheet.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_data[(row, col)] = top_left_val

    # 2. Extract all non-empty cells to find the bounding box
    cells_dict = {}
    min_r, max_r, min_c, max_c = float('inf'), -1, float('inf'), -1
    
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            # Use merged value if cell is in a merged range, otherwise use cell value
            val = merged_data.get((row, col), cell.value)
            
            if val is not None and str(val).strip() != "":
                cells_dict[(row, col)] = val
                min_r = min(min_r, row)
                max_r = max(max_r, row)
                min_c = min(min_c, col)
                max_c = max(max_c, col)

    if not cells_dict:
        return []

    # 3. Robust Header Row Detection
    header_start_row = None
    header_depth_detected = 1
    final_headers = []
    
    for r in range(min_r, min(min_r + 20, max_r + 1)):
        non_null_cells = 0
        text_cells = 0
        numeric_cells = 0
        
        for c in range(min_c, max_c + 1):
            val = cells_dict.get((r, c))
            if val is not None and str(val).strip() != "":
                non_null_cells += 1
                if isinstance(val, (int, float, complex)) and not isinstance(val, bool):
                    numeric_cells += 1
                else:
                    text_cells += 1
                    
        # Check condition: non_null_cells >= 5, text >= 70%, numeric <= 30%
        # Fallback to len of columns if table is smaller than 5 columns
        min_required = min(5, max_c - min_c + 1)
        if non_null_cells >= min_required:
            text_ratio = text_cells / non_null_cells
            numeric_ratio = numeric_cells / non_null_cells
            
            if text_ratio >= 0.70 and numeric_ratio <= 0.30:
                best_d = 1
                max_score = -1
                best_headers = []
                
                # Check up to 4 rows deep to find optimal combined header depth
                for d in range(1, min(5, max_r - r + 2)):
                    headers_at_d = []
                    header_block = {}
                    
                    # Manual horizontal forward fill for top-level merged headers
                    # (this fixes implicitly merged top-headers without explicit styling)
                    for hr in range(r, r + d):
                        last_val = ""
                        for c in range(min_c, max_c + 1):
                            val = cells_dict.get((hr, c))
                            if val is not None and str(val).strip() != "":
                                last_val = str(val).strip()
                            
                            # Only forward fill the uppermost grouping headers, keep bottom row exact
                            if hr < r + d - 1:
                                header_block[(hr, c)] = last_val
                            else:
                                header_block[(hr, c)] = str(val).strip() if val is not None else ""
                                
                    for c in range(min_c, max_c + 1):
                        col_parts = []
                        for hr in range(r, r + d):
                            part = header_block.get((hr, c), "")
                            if part and (not col_parts or col_parts[-1] != part):
                                col_parts.append(part)
                                
                        # Keep chronological order: Top-level header first, then specific sub-headers
                        # Example: 'Ld Drawing Release' + 'Plan' -> ld_drawing_release_plan
                        
                        header_name = " ".join(col_parts).lower().strip()
                        # Replace dots and hyphens with spaces first to become standard underscores
                        header_name = re.sub(r'[.\-]', ' ', header_name)
                        # Remove anything that isn't alphanumeric, underscore, or space
                        header_name = re.sub(r'[^a-z0-9_\s]', '', header_name)
                        # Consolidate spaces into single underscore
                        header_name = re.sub(r'\s+', '_', header_name).strip('_')
                        
                        if not header_name:
                            header_name = f"column_{c}"
                            
                        headers_at_d.append(header_name)
                        
                    unique_count = len(set(headers_at_d))
                    generic_count = sum(1 for h in headers_at_d if h.startswith("column_"))
                    
                    # Score function maximizes the number of uniquely identifiable columns
                    # and penalizes any empty placeholders
                    score = unique_count - (generic_count * 0.5)
                    
                    if score > max_score:
                        max_score = score
                        best_d = d
                        best_headers = headers_at_d
                
                has_generic = any(h.startswith("column_") for h in best_headers)
                
                # If the combined header set is completely clean, save it and proceed!
                if not has_generic:
                    header_start_row = r
                    header_depth_detected = best_d
                    final_headers = best_headers
                    break

    if header_start_row is None:
        raise ValueError("Could not detect a valid header row. Ensure your sheet has a header with at least 5 text columns and no generic placeholders.")
        
    header_rows_end = header_start_row + header_depth_detected - 1

    # Resolve duplicate column names
    unique_headers = []
    header_counts = {}
    for h in final_headers:
        if h in header_counts:
            header_counts[h] += 1
            unique_headers.append(f"{h}_{header_counts[h]}")
        else:
            header_counts[h] = 0
            unique_headers.append(h)

    # 5. Extract rows and infer basic data types gracefully
    data = []
    for r in range(header_rows_end + 1, max_r + 1):
        row_data = {}
        is_empty = True
        for idx, c in enumerate(range(min_c, max_c + 1)):
            val = cells_dict.get((r, c))
            
            # Openpyxl parses dates to datetime objects natively. JSON can't serialize them.
            if isinstance(val, (datetime.date, datetime.datetime)):
                val = val.isoformat()
                
            if val is not None:
                is_empty = False
                
            row_data[unique_headers[idx]] = val
            
        # Avoid appending completely null rows
        if not is_empty:
            data.append(row_data)

    return data


async def create_dynamic_table(conn, base_name: str, data: list):
    if not data:
        return None
    safe_name = re.sub(r'[^a-zA-Z0-9_]', '_', base_name).lower()
    if not safe_name or not safe_name[0].isalpha():
        safe_name = 'tbl_' + safe_name
    table_name = f"{safe_name[:40]}_{int(time.time())}"
    
    columns = list(data[0].keys())
    
    # Safe column names mapping (avoid quotes issues)
    safe_columns = []
    for col in columns:
        col_name = str(col).replace('"', '""')
        safe_columns.append(f'"{col_name}" TEXT')
        
    cols_def = ", ".join(safe_columns)
    create_stmt = f'CREATE TABLE {table_name} (id SERIAL PRIMARY KEY, {cols_def})'
    await conn.execute(create_stmt)
    
    insert_columns = []
    for col in columns:
        col_safe = str(col).replace('"', '""')
        insert_columns.append(f'"{col_safe}"')
        
    insert_cols = ", ".join(insert_columns)
    placeholders = ", ".join([f"${i+1}" for i in range(len(columns))])
    insert_stmt = f"INSERT INTO {table_name} ({insert_cols}) VALUES ({placeholders})"
    
    values = []
    for row in data:
        values.append([str(row.get(col, "")) if row.get(col) is not None else None for col in columns])
        
    await conn.executemany(insert_stmt, values)
    return table_name

@app.post("/upload")
async def upload_excel(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file format. Upload .xlsx or .xls")
        
    contents = await file.read()
    
    try:
        processed_data = process_excel(contents)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing excel: {str(e)}")
        
    if not processed_data:
        raise HTTPException(status_code=400, detail="No readable data found.")
        
    if app.state.pool is None:
        raise HTTPException(status_code=500, detail="Database connection not available")

    # Store in new dynamic table
    try:
        async with app.state.pool.acquire() as conn:
            base_filename = file.filename.rsplit('.', 1)[0]
            table_name = await create_dynamic_table(conn, base_filename, processed_data)
            await conn.execute("INSERT INTO file_registry (filename, table_name) VALUES ($1, $2)", file.filename, table_name)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database insertion failed: {str(e)}")
        
    return {
        "message": "File processed and saved as database table.", 
        "table_name": table_name,
        "data": processed_data,
        "rows_processed": len(processed_data),
        "columns_detected": len(processed_data[0].keys())
    }

if __name__ == "__main__":
    uvicorn.run("main_ingestion_only:app", host="0.0.0.0", port=8000, reload=True)
